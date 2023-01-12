# -*- coding: utf-8 -*-
import os
import re
from datetime import datetime

import openpyxl
from django.conf import settings

from dvadmin.utils.validator import CustomValidationError


def import_to_data(file_url, field_data, m2m_fields=None):
    """
    读取导入的excel文件
    :param file_url:
    :param field_data: 首行数据源
    :param m2m_fields: 多对多字段
    :return:
    """
    # 读取excel 文件
    file_path_dir = os.path.join(settings.BASE_DIR, file_url)
    workbook = openpyxl.load_workbook(file_path_dir)
    table = workbook[workbook.sheetnames[0]]
    theader = tuple(table.values)[0] #Excel的表头
    is_update = '更新主键(勿改)' in theader #是否导入更新
    if is_update is False: #不是更新时,删除id列
        field_data.pop('id')
    # 获取参数映射
    validation_data_dict = {}
    for key, value in field_data.items():
        if isinstance(value, dict):
            choices = value.get("choices", {})
            data_dict = {}
            if choices.get("data"):
                for k, v in choices.get("data").items():
                    data_dict[k] = v
            elif choices.get("queryset") and choices.get("values_name"):
                data_list = choices.get("queryset").values(choices.get("values_name"), "id")
                for ele in data_list:
                    data_dict[ele.get(choices.get("values_name"))] = ele.get("id")
            else:
                continue
            validation_data_dict[key] = data_dict
    # 创建一个空列表，存储Excel的数据
    tables = []
    for i, row in enumerate(range(table.max_row)):
        if i == 0:
            continue
        array = {}
        for index, item in enumerate(field_data.items()):
            items = list(item)
            key = items[0]
            values = items[1]
            value_type = 'str'
            if isinstance(values, dict):
                value_type = values.get('type','str')
            cell_value = table.cell(row=row + 1, column=index + 2).value
            if cell_value is None or cell_value=='':
                continue
            elif value_type == 'date':
                print(61, datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S').date())
                try:
                    cell_value = datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S').date()
                except:
                    raise CustomValidationError('日期格式不正确')
            elif value_type == 'datetime':
                cell_value = datetime.strptime(str(cell_value), '%Y-%m-%d %H:%M:%S')
            else:
            # 由于excel导入数字类型后，会出现数字加 .0 的，进行处理
                if type(cell_value) is float and str(cell_value).split(".")[1] == "0":
                    cell_value = int(str(cell_value).split(".")[0])
                elif type(cell_value) is str:
                    cell_value = cell_value.strip(" \t\n\r")
            if key in validation_data_dict:
                array[key] = validation_data_dict.get(key, {}).get(cell_value, None)
                if key in m2m_fields:
                    array[key] = list(
                        filter(
                            lambda x: x,
                            [
                                validation_data_dict.get(key, {}).get(value, None)
                                for value in re.split(r"[，；：|.,;:\s]\s*", cell_value)
                            ],
                        )
                    )
            else:
                array[key] = cell_value
        tables.append(array)
    return tables
