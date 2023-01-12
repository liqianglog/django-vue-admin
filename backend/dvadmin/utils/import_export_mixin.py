# -*- coding: utf-8 -*-
from urllib.parse import quote

from django.db import transaction
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework.decorators import action
from rest_framework.request import Request

from dvadmin.utils.import_export import import_to_data
from dvadmin.utils.json_response import DetailResponse
from dvadmin.utils.request_util import get_verbose_name


class ImportSerializerMixin:
    """
    自定义导入模板、导入功能
    """

    # 导入字段
    import_field_dict = {}
    # 导入序列化器
    import_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get','post'],detail=False)
    @transaction.atomic  # Django 事务,防止出错
    def import_data(self, request: Request, *args, **kwargs):
        """
        导入模板
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        assert self.import_field_dict, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        # 导出模板
        if request.method == "GET":
            # 示例数据
            queryset = self.filter_queryset(self.get_queryset())
            # 导出excel 表
            response = HttpResponse(content_type="application/msexcel")
            response["Access-Control-Expose-Headers"] = f"Content-Disposition"
            response[
                "Content-Disposition"
            ] = f'attachment;filename={quote(str(f"导入{get_verbose_name(queryset)}模板.xlsx"))}'
            wb = Workbook()
            ws1 = wb.create_sheet("data", 1)
            ws1.sheet_state = "hidden"
            ws = wb.active
            row = get_column_letter(len(self.import_field_dict) + 1)
            column = 10
            header_data = [
                "序号",
            ]
            validation_data_dict = {}
            for index, ele in enumerate(self.import_field_dict.values()):
                if isinstance(ele, dict):
                    header_data.append(ele.get("title"))
                    choices = ele.get("choices", {})
                    if choices.get("data"):
                        data_list = []
                        data_list.extend(choices.get("data").keys())
                        validation_data_dict[ele.get("title")] = data_list
                    elif choices.get("queryset") and choices.get("values_name"):
                        data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                        validation_data_dict[ele.get("title")] = list(data_list)
                    else:
                        continue
                    column_letter = get_column_letter(len(validation_data_dict))
                    dv = DataValidation(
                        type="list",
                        formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[ele.get('title')]) + 1}",
                        allow_blank=True,
                    )
                    ws.add_data_validation(dv)
                    dv.add(f"{get_column_letter(index + 2)}2:{get_column_letter(index + 2)}1048576")
                else:
                    header_data.append(ele)
            # 添加数据列
            ws1.append(list(validation_data_dict.keys()))
            for index, validation_data in enumerate(validation_data_dict.values()):
                for inx, ele in enumerate(validation_data):
                    ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
            # 插入导出模板正式数据
            df_len_max = [self.get_string_len(ele) for ele in header_data]
            ws.append(header_data)
            # 　更新列宽
            for index, width in enumerate(df_len_max):
                ws.column_dimensions[get_column_letter(index + 1)].width = width
            tab = Table(displayName="Table1", ref=f"A1:{row}{column}")  # 名称管理器
            style = TableStyleInfo(
                name="TableStyleLight11",
                showFirstColumn=True,
                showLastColumn=True,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            wb.save(response)
            return response
        else:
            # 从excel中组织对应的数据结构，然后使用序列化器保存
            queryset = self.filter_queryset(self.get_queryset())
            # 获取多对多字段
            m2m_fields = [
                ele.name
                for ele in queryset.model._meta.get_fields()
                if hasattr(ele, "many_to_many") and ele.many_to_many == True
            ]
            import_field_dict = {'id':'更新主键(勿改)',**self.import_field_dict}
            data = import_to_data(request.data.get("url"), import_field_dict, m2m_fields)
            for ele in data:
                filter_dic = {'id':ele.get('id')}
                instance = filter_dic and queryset.filter(**filter_dic).first()
                # print(156,ele)
                serializer = self.import_serializer_class(instance, data=ele, request=request)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            return DetailResponse(msg=f"导入成功！")

    @action(methods=['get'],detail=False)
    def update_template(self,request):
        queryset = self.filter_queryset(self.get_queryset())
        assert self.import_field_dict, "'%s' 请配置对应的导入模板字段。" % self.__class__.__name__
        assert self.import_serializer_class, "'%s' 请配置对应的导入序列化器。" % self.__class__.__name__
        data = self.import_serializer_class(queryset, many=True, request=request).data
        # 导出excel 表
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active
        import_field_dict = {}
        header_data = ["序号","更新主键(勿改)"]
        hidden_header = ["#","id"]
        #----设置选项----
        validation_data_dict = {}
        for index, item in enumerate(self.import_field_dict.items()):
            items = list(item)
            key = items[0]
            value = items[1]
            if isinstance(value, dict):
                header_data.append(value.get("title"))
                hidden_header.append(value.get('display'))
                choices = value.get("choices", {})
                if choices.get("data"):
                    data_list = []
                    data_list.extend(choices.get("data").keys())
                    validation_data_dict[value.get("title")] = data_list
                elif choices.get("queryset") and choices.get("values_name"):
                    data_list = choices.get("queryset").values_list(choices.get("values_name"), flat=True)
                    validation_data_dict[value.get("title")] = list(data_list)
                else:
                    continue
                column_letter = get_column_letter(len(validation_data_dict))
                dv = DataValidation(
                    type="list",
                    formula1=f"{quote_sheetname('data')}!${column_letter}$2:${column_letter}${len(validation_data_dict[value.get('title')]) + 1}",
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(index + 3)}2:{get_column_letter(index + 3)}1048576")
            else:
                header_data.append(value)
                hidden_header.append(key)
        # 添加数据列
        ws1.append(list(validation_data_dict.keys()))
        for index, validation_data in enumerate(validation_data_dict.values()):
            for inx, ele in enumerate(validation_data):
                ws1[f"{get_column_letter(index + 1)}{inx + 2}"] = ele
        #--------
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(hidden_header) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        elif isinstance(val,list):
                            results_list.append(str(val))
                        else:
                            results_list.append(val)
                        # 计算最大列宽度
                        if isinstance(val,str):
                            result_column_width = self.get_string_len(val)
                            if h_index != 0 and result_column_width > df_len_max[h_index]:
                                df_len_max[h_index] = result_column_width
            ws.append([index+1,*results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response


class ExportSerializerMixin:
    """
    自定义导出功能
    """

    # 导出字段
    export_field_label = []
    # 导出序列化器
    export_serializer_class = None
    # 表格表头最大宽度，默认50个字符
    export_column_width = 50

    def is_number(self,num):
        try:
            float(num)
            return True
        except ValueError:
            pass

        try:
            import unicodedata
            unicodedata.numeric(num)
            return True
        except (TypeError, ValueError):
            pass
        return False

    def get_string_len(self, string):
        """
        获取字符串最大长度
        :param string:
        :return:
        """
        length = 4
        if string is None:
            return length
        if self.is_number(string):
            return length
        for char in string:
            length += 2.1 if ord(char) > 256 else 1
        return round(length, 1) if length <= self.export_column_width else self.export_column_width

    @action(methods=['get'],detail=False)
    def export_data(self, request: Request, *args, **kwargs):
        """
        导出功能
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        queryset = self.filter_queryset(self.get_queryset())
        assert self.export_field_label, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        assert self.export_serializer_class, "'%s' 请配置对应的导出序列化器。" % self.__class__.__name__
        data = self.export_serializer_class(queryset, many=True, request=request).data
        # 导出excel 表
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["序号", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key,val in results.items():
                    if key == h_item:
                        if val is None or val=="":
                            results_list.append("")
                        else:
                            results_list.append(val)
                        # 计算最大列宽度
                        result_column_width = self.get_string_len(val)
                        if h_index !=0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response
